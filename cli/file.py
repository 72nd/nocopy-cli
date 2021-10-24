"""
Everything related to file in- and output.
"""

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from abc import (
    ABC,
    abstractmethod,
    abstractclassmethod,
    abstractstaticmethod,
)
import csv
import io
import json
import logging
import math
from pathlib import Path
import sys
from typing import Any, Dict, List, Optional
import yaml


class StdInNotSupported(Exception):
    """
    Exception raised when user tries to load data into the application using
    the standard input (stdin) for a `File` implementation which doesn't
    support this. Also see `File.__supports_std` for more information on that
    matter.
    """

    def __init__(self, file: "File"):
        super().__init__(
            f"{file.format_name()} does not support the input of data using "
            "stdin. Please specify a input path."
        )


class StdOutNotSupported(Exception):
    """
    Exception raised when user tries to load data into the application using
    the standard output (stdout) for a `File` implementation which doesn't
    support this. Also see `File.__supports_std` for more information on that
    matter.
    """

    def __init__(self, file: "File"):
        super().__init__(
            f"{file.format_name()} does not support the input of data using "
            "stdout. Please specify a output path."
        )


class FormatUnknown(Exception):
    """
    Raised when a unknown format string is given.
    """

    def __init__(self, value: str):
        super().__init__(
            f"{value} is not supported by this application",
        )


class FormatNotAscertainable(Exception):
    """
    Raised when the `File` type couldn't be determined for a file path.
    """

    def __init__(self, file: Path):
        super().__init__(
            f"file '{file}' with extension {file.suffix} is not supported"
        )


class DifferentInOutFormats(Exception):
    """
    Raised when the given input path has another file suffix than the output
    path. This is a hypothetical edge case which has no real-life implication
    as the application currently never uses a input and a output file at the
    same time.
    """

    def __init__(self):
        super().__init__(
            "file extension of input file doesn't match it's counterpart of "
            "the output file"
        )


class File(ABC):
    """
    Provides the in- and output of data into and out of nocopy-cli.

    A file can also be a stdin or out. This is choosen if no input or output
    file was given by the user. The children classes of `File` implement the
    functionality for a certain data type (like JSON).
    """

    input_path: Optional[Path]
    """Optional path to input file."""
    output_path: Optional[Path]
    """Optional path to input file."""
    level_nested: bool
    """Level out nested structures."""

    def __init__(
        self,
        input_path: Optional[Path],
        output_path: Optional[Path],
        level_nested: bool,
    ):
        self.input_path = input_path
        self.output_path = output_path
        self.level_nested = level_nested

    def load(self) -> Dict[str, Any]:
        """Load the data in the file and returns the data as a dict."""
        if self.input_path is None:
            if not self.supports_std():
                raise StdInNotSupported(self)
            return self.parse(io.BytesIO(self.__read_stdin()))

        with open(self.input_path, "rb") as file:
            return self.parse(file)

    def save(self, data: Dict[str, Any]):
        """Saves/outputs the data to the file/stdout."""
        data = self.__level(data)
        if self.output_path is None:
            if not self.supports_std():
                raise StdOutNotSupported(self)
            print(self.dump(data).decode("utf-8"))
        else:
            with open(self.output_path, "wb") as file:
                file.write(self.dump(data))

    @staticmethod
    def __read_stdin() -> str:
        rsl = ""
        try:
            for line in sys.stdin:
                rsl += line
        except KeyboardInterrupt:
            sys.stdout.flush()
        return rsl

    @abstractclassmethod
    def format_name(cls) -> str:
        """
        Returns the name of the format supported by the given implementation of
        `File`. This is value is also used by the user interface to name this
        format.
        """
        pass

    @abstractclassmethod
    def file_extensions(cls) -> List[str]:
        """
        Returns a list of the valid file extensions for this file type with a
        leading dot.
        """
        pass

    @abstractclassmethod
    def supports_std(self) -> bool:
        """
        States whether the data type can be in-/outputted to/from the
        application via stdin and stdout. Motivation: Some file formats do not
        really make sense to be piped between multiple application in a shell
        like Excel's xlsx format.
        """
        pass

    @abstractstaticmethod
    def parse(raw: io.BufferedReader) -> Dict[str, Any]:
        """
        Parses a given file (opened in binary mode) and returns it's structure
        as a dictionary.
        """
        pass

    @abstractmethod
    def dump(self, data: Dict[str, Any]) -> bytes:
        """
        Return a bytes literal of the dumped data structured.
        """
        pass

    def __level(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not self.level_nested:
            return data
        for entry in data:
            rsl = {}
            for field in entry:
                if not isinstance(entry[field], dict):
                    continue
                sub_rsl = {}
                for sub_field in entry[field]:
                    sub_rsl[f"{field}_{sub_field}"] = entry[field][sub_field]
                rsl[field] = sub_rsl
            for r in rsl:
                del entry[r]
                entry.update(rsl[r])

        return data


class Json(File):
    """JSON implementation for `File`."""

    def __init__(self, *args):
        super().__init__(*args)

    @classmethod
    def format_name(cls) -> str:
        return "json"

    @classmethod
    def file_extensions(cls) -> List[str]:
        return [".json"]

    @classmethod
    def supports_std(self) -> bool:
        return True

    @staticmethod
    def parse(raw: io.BufferedReader) -> Dict[str, Any]:
        return json.load(raw)

    def dump(self, data: Dict[str, Any]) -> bytes:
        return json.dumps(data).encode("utf-8")


class Yaml(File):
    """YAML implementation for `File`."""

    def __init__(self, *args):
        super().__init__(*args)

    @classmethod
    def format_name(cls) -> str:
        return "yaml"

    @classmethod
    def file_extensions(cls) -> List[str]:
        return [".yaml", ".yml"]

    @classmethod
    def supports_std(self) -> bool:
        return True

    @staticmethod
    def parse(raw: io.BufferedReader) -> Dict[str, Any]:
        return yaml.load(raw, Loader=yaml.FullLoader)

    def dump(self, data: Dict[str, Any]) -> bytes:
        return yaml.dump(data).encode("utf-8")


class Csv(File):
    """CSV implementation for `File`."""

    only_header: bool = False
    """Only write the header."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args)

        if "only_header" in kwargs:
            self.only_header = kwargs["only_header"]
        if "level_nested" in kwargs:
            self.level_nested = kwargs["level_nested"]

    @classmethod
    def format_name(cls) -> str:
        return "csv"

    @classmethod
    def file_extensions(cls) -> List[str]:
        return [".csv"]

    @classmethod
    def supports_std(self) -> bool:
        return True

    @staticmethod
    def parse(raw: io.BufferedReader) -> Dict[str, Any]:
        data = csv.DictReader(io.StringIO(raw.read().decode("utf-8")))
        rsl = []
        for entry in data:
            for field in entry:
                if entry[field] == "":
                    entry[field] = None
            rsl.append(entry)
        return rsl

    def dump(self, data: List[Dict[str, Any]]) -> bytes:
        buffer = io.StringIO()
        writer = csv.DictWriter(
            buffer,
            fieldnames=data[0].keys(),
            quotechar='"',
        )
        writer.writeheader()
        if not self.only_header:
            for entry in data:
                writer.writerow(entry)
        return buffer.getvalue().encode("utf-8")


class Xlsx(File):
    """
    The Office Open XML Workbook implementation for `File`. Used by Excel.
    """

    freeze_at: Optional[str] = None
    """Defines the cell where the table should be freezed at."""
    header_color = "00A9A9A9"
    """Defines the background color of the header row."""
    header_font: Font = Font(
        name="Calibri",
        bold=True,
    )
    font: Font = Font(
        name="Calibri",
        size=10,
    )
    """Font format for the header row."""
    side: Side = Side(border_style="thin", color="000000")
    """Line style of the cells."""
    alternating_row_color: str = "00D3D3D3"
    """Background of the alternating rows."""

    __longest_entry_temp: Optional[Dict[str, Any]] = None


    def __init__(self, *args, **kwargs):
        super().__init__(*args)
        if "freeze_at" in kwargs:
            self.freeze_at = kwargs["freeze_at"]

    @classmethod
    def format_name(cls) -> str:
        return "xlsx"

    @classmethod
    def file_extensions(cls) -> List[str]:
        return [".xlsx"]

    @classmethod
    def supports_std(self) -> bool:
        return True

    @staticmethod
    def parse(raw: io.BufferedReader) -> Dict[str, Any]:
        raise NotImplementedError()

    def dump(self, data: List[Dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        self.__worksheet_insert_header(ws, data)
        self.__worksheet_insert_content(ws, data)
        self.__freeze_cells(ws)
        self.__apply_apperance(wb, ws, data)
        self.__apply_conditional(ws, data)
        self.__apply_data_table(ws, data)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def __worksheet_insert_header(
        self,
        ws: Worksheet,
        data: List[Dict[str, Any]],
    ) -> None:
        """Sets and formats the header of the Worksheet."""
        header_row = ws.row_dimensions[1]
        ws.append([name for name in self.__longest_data_entry(data)])

    def __worksheet_insert_content(
        self,
        ws: Worksheet,
        data: List[Dict[str, Any]],
    ) -> None:
        """Inserts the actual data into the Worksheet."""
        for row in data:
            ws.append([str(cell) for cell in row.values()])

    def __freeze_cells(self, ws: Worksheet) -> None:
        """Freezes the cell at the given cell from the config."""
        if self.freeze_at is not None:
            ws.freeze_panes = ws[self.freeze_at]

    def __apply_apperance(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: List[Dict[str, Any]],
    ) -> None:
        border = Border(
            left=self.side,
            right=self.side,
            top=self.side,
            bottom=self.side,
            vertical=self.side,
            horizontal=self.side,
        )
        header_style = NamedStyle(name="Header")
        header_style.alignment = Alignment(vertical="center")
        header_style.border = border
        header_style.fill = PatternFill("solid", fgColor=self.header_color)
        header_style.font = self.header_font

        default_style = NamedStyle(name="Default")
        default_style.alignment = Alignment(vertical="center")
        default_style.border = border
        default_style.font = self.font

        widths = self.__calc_column_widths(data)
        for col in range(1, ws.max_column+1):
            column_letter = get_column_letter(col)
            column = ws.column_dimensions[column_letter]
            column.width = widths[col-1]
            if widths[col-1] == 80:
                column.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                    shrink_to_fit=True,
                )
            for row in range(1, len(data)+1):
                if row == 1:
                    style = header_style
                else:
                    style = default_style
                ws[f"{column_letter}{row}"].style = style

    def __apply_conditional(
        self,
        ws: Worksheet,
        data: List[Dict[str, Any]],
    ) -> None:
        alternating_fill = fill=PatternFill(
            "solid",
            bgColor=self.alternating_row_color,
        )
        alternating_style = DifferentialStyle(fill=alternating_fill)
        alternating_rule = FormulaRule(
            fill=alternating_fill,
            formula=["ISODD(ROW())"],
        )
        ws.conditional_formatting.add(
            f"A2:{self.__max_cell(data)}",
            alternating_rule,
        )

    def __apply_data_table(
        self,
        ws: Worksheet,
        data: List[Dict[str, Any]],
    ) -> None:
        """Configures the data table."""
        table = Table(
            displayName="Table",
            ref=self.__dimensions(data)
        )
        ws.add_table(table)

    def __longest_data_entry(
        self,
        data: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        if self.__longest_entry_temp is not None:
            return self.__longest_entry_temp
        longest = 0
        for i in range(0, len(data)):
            if len(data[i]) > len(data[longest]):
                longest = i
        self.__longest_entry_temp = data[longest]
        return data[longest]

    def __calc_column_widths(
        self,
        data: List[Dict[str, Any]],
    ) -> None:
        """Calculates an approximative width for each column."""
        keys = [key for key in self.__longest_data_entry(data)]

        widths_per_key: Dict[str, int] = {}
        for key in keys:
            widths_per_key[key] = 0

        header_data = data[0]
        for key in keys:
            header_data[key] = key
        data.append(header_data)

        for entry in data:
            entry_dict = entry
            for key in keys:
                if key not in entry_dict or not isinstance(entry_dict[key], str):
                    continue
                if (length := len(entry_dict[key])) > widths_per_key[key]:
                    widths_per_key[key] = length
        rsl: List[int] = []
        for value in widths_per_key.values():
            width = math.ceil(value*0.93)
            if width > 80:
                rsl.append(80)
            elif width > 5:
                rsl.append(width)
            else:
                rsl.append(5)
        return rsl

    def __max_cell(self, data: List[Dict[str, Any]]) -> None:
        """Returns the coordinate of the most outer cell."""
        column = get_column_letter(len(self.__longest_data_entry(data)))
        row = len(data)
        return f"{column}{row}"

    def __dimensions(self, data: List[Dict[str, Any]]) -> None:
        """Returns the dimension of the used space."""
        return f"A1:{self.__max_cell(data)}"


def file(
    format_option: Optional[str] = None,
    input_path: Optional[Path] = None,
    output_path: Optional[Path] = None,
    level_nested: bool = False,
    **kwargs,
) -> File:
    """
    Factory method for `File`.

    Returns the appropriate `File` implementation based on the given
    format_option.
    """
    if format_option is None and input_path is None and output_path is None:
        logging.debug("no format specified fall back to default YAML")
        return Yaml(input_path, output_path, level_nested)

    if format_option is not None:
        format_option = format_option.lower()
    if format_option == Json.format_name():
        return Json(input_path, output_path, level_nested)
    elif format_option == Yaml.format_name():
        return Yaml(input_path, output_path, level_nested)
    elif format_option == Csv.format_name():
        return Csv(input_path, output_path, level_nested)
    elif format_option == Xlsx.format_name():
        return Xlsx(input_path, output_path, level_nested, **kwargs)
    elif format_option is not None:
        raise FormatUnknown(format_option)

    if input_path is not None and output_path is not None and \
            input_path.suffix.lower() != output_path.suffix.lower():
        raise DifferentInOutFormats()

    if input_path is not None:
        path = input_path
    else:
        path = output_path
    suffix = path.suffix.lower()

    if suffix in Json.file_extensions():
        logging.debug("type assessed as JSON by file extension")
        return Json(input_path, output_path, level_nested)
    elif suffix in Yaml.file_extensions():
        logging.debug("type assessed as YAML by file extension")
        return Yaml(input_path, output_path, level_nested)
    elif suffix in Csv.file_extensions():
        logging.debug("type assessed as CSV by file extension")
        return Csv(input_path, output_path, level_nested)
    elif suffix in Xlsx.file_extensions():
        logging.debug("type assessed as XLSX by file extension")
        return Xlsx(input_path, output_path, level_nested, **kwargs)
    raise FormatNotAscertainable(path)
